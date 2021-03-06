 # -*- coding: utf-8 -*-
from flask_sqlalchemy import SQLAlchemy
from openpyxl import load_workbook

from anyway.core.utils import Utils
from anyway.common.models.cbs_models import RoadSegments


def _iter_rows(filename):
    workbook = load_workbook(filename, read_only=True)
    sheet = workbook[u"tv_ktaim"]
    rows = sheet.rows
    first_row = next(rows)
    headers = [u'mezahe_keta', u'kvish', u'keta', u'km_me', u'shem_km_me', u'ad_km', u'shem_km_ad']
    assert [cell.value for cell in first_row] == headers
    for row in rows:
        segment_id = row[0].value

        # In order to ignore empty lines
        if not segment_id:
            continue
        road = row[1].value
        segment = row[2].value
        from_km = row[3].value
        from_name = row[4].value
        to_km = row[5].value
        to_name = row[6].value

        yield {
            'segment_id': segment_id,
            'road': road,
            'segment': segment,
            'from_km': from_km,
            'from_name': from_name,
            'to_km': to_km,
            'to_name': to_name,
        }


def parse(filename):
    from anyway import db

    for batch in Utils.batch_iterator(_iter_rows(filename), batch_size=50):
        db.session.bulk_insert_mappings(RoadSegments, batch)
        db.session.commit()
